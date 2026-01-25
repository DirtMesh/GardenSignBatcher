from __future__ import annotations

from pathlib import Path
from typing import List, Tuple, Optional

from openpyxl import load_workbook

from .models import InputRow, ValidationIssue


def list_sheets(xlsx_path: Path) -> list[str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def read_rows_from_sheet(
    xlsx_path: Path,
    sheet_name: str,
) -> Tuple[List[InputRow], List[ValidationIssue]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        # Assume headers are in row 1
        headers = []
        for cell in ws[1]:
            headers.append((str(cell.value).strip() if cell.value is not None else "").lower())

        def col_index(name: str) -> Optional[int]:
            try:
                return headers.index(name)
            except ValueError:
                return None

        crop_idx = col_index("crop_name")
        cult_idx = col_index("cultivar")

        issues: list[ValidationIssue] = []
        if crop_idx is None:
            issues.append(ValidationIssue(row_num=1, reason="Missing required column: crop_name"))
            return [], issues

        rows: list[InputRow] = []
        # Iterate from row 2 onward
        for r in range(2, ws.max_row + 1):
            crop = ws.cell(row=r, column=crop_idx + 1).value
            cult = ws.cell(row=r, column=(cult_idx + 1) if cult_idx is not None else 1).value if cult_idx is not None else ""

            crop_s = (str(crop).strip() if crop is not None else "")
            cult_s = (str(cult).strip() if cult is not None else "")

            if not crop_s:
                issues.append(ValidationIssue(row_num=r, reason="Missing crop_name"))
                continue

            rows.append(InputRow(row_num=r, crop_name=crop_s, cultivar=cult_s))

        return rows, issues
    finally:
        wb.close()
